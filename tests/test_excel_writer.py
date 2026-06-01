"""Tests for excel_writer.py — append_row, remove_last_row, read_all_categories."""

import pytest
import openpyxl
from datetime import datetime

from expense_agent.excel_writer import append_row, remove_last_row, read_all_categories, ExcelError


# ---------- append_row ----------

class TestAppendRow:
    def _make_txn(self, **overrides):
        txn = {"date": "02/21/2026", "item": "Test Store", "category": "Misc",
               "subcategory": "Misc", "amount": 5.25}
        txn.update(overrides)
        return txn

    def test_writes_correct_columns(self, config):
        append_row(config, self._make_txn())
        wb = openpyxl.load_workbook(config["excel_path"])
        ws = wb["Daily Expenses"]
        # Row 2 is first data row (row 1 is header from fixture)
        assert ws.cell(row=2, column=1).value == 2026       # Year
        assert ws.cell(row=2, column=2).value == "Feb"       # Month
        assert ws.cell(row=2, column=4).value == 5.25        # Amount
        assert ws.cell(row=2, column=5).value == "Misc"      # Category
        assert ws.cell(row=2, column=6).value == "Misc"      # Subcategory
        assert ws.cell(row=2, column=7).value == "Test Store" # Item
        wb.close()

    def test_multiple_rows(self, config):
        append_row(config, self._make_txn(item="Store A"))
        append_row(config, self._make_txn(item="Store B"))
        wb = openpyxl.load_workbook(config["excel_path"])
        ws = wb["Daily Expenses"]
        assert ws.cell(row=2, column=7).value == "Store A"
        assert ws.cell(row=3, column=7).value == "Store B"
        wb.close()

    def test_file_not_found_raises(self, tmp_path):
        cfg = {"excel_path": str(tmp_path / "nonexistent.xlsx"), "sheet_name": "Sheet1"}
        with pytest.raises(ExcelError):
            append_row(cfg, self._make_txn())

    def test_sheet_not_found_raises(self, config):
        config["sheet_name"] = "Wrong Sheet"
        with pytest.raises(ExcelError):
            append_row(config, self._make_txn())

    def test_date_parsing(self, config):
        append_row(config, self._make_txn(date="03/15/2026"))
        wb = openpyxl.load_workbook(config["excel_path"])
        ws = wb["Daily Expenses"]
        date_val = ws.cell(row=2, column=3).value
        assert isinstance(date_val, datetime)
        assert date_val.month == 3
        assert date_val.day == 15
        wb.close()


# ---------- remove_last_row ----------

class TestRemoveLastRow:
    def test_returns_transaction(self, config):
        txn = {"date": "02/21/2026", "item": "Test", "category": "Misc",
               "subcategory": "Misc", "amount": 5.25}
        append_row(config, txn)
        removed = remove_last_row(config)
        assert removed is not None
        assert removed["item"] == "Test"
        assert removed["subcategory"] == "Misc"
        assert removed["amount"] == 5.25

    def test_empty_returns_none(self, config):
        result = remove_last_row(config)
        assert result is None

    def test_actually_deletes(self, config):
        txn = {"date": "02/21/2026", "item": "Test", "category": "Misc",
               "subcategory": "Misc", "amount": 5.25}
        append_row(config, txn)
        remove_last_row(config)
        wb = openpyxl.load_workbook(config["excel_path"])
        ws = wb["Daily Expenses"]
        assert ws.max_row == 1  # only header remains
        wb.close()

    def test_file_not_found_raises(self, tmp_path):
        cfg = {"excel_path": str(tmp_path / "nope.xlsx"), "sheet_name": "Sheet1"}
        with pytest.raises(ExcelError):
            remove_last_row(cfg)


# ---------- read_all_categories ----------

class TestReadAllCategories:
    def test_returns_mapping(self, config):
        txn = {"date": "02/21/2026", "item": "Coffee Shop", "category": "Food & Dining",
               "subcategory": "Dining", "amount": 4.50}
        append_row(config, txn)
        mapping = read_all_categories(config)
        assert mapping["Coffee Shop"] == "Dining"

    def test_missing_file(self, tmp_path):
        cfg = {"excel_path": str(tmp_path / "nope.xlsx"), "sheet_name": "Sheet1"}
        assert read_all_categories(cfg) == {}

    def test_missing_sheet(self, config):
        config["sheet_name"] = "Nonexistent"
        assert read_all_categories(config) == {}

    def test_skips_empty_cells(self, tmp_path):
        # Create a workbook with no header data in cols 6-7
        excel_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Expenses"
        # Row 1: header with no subcategory/item
        ws.cell(row=1, column=1, value="Year")
        # Row 2: data row with item cleared
        ws.cell(row=2, column=6, value="Misc")
        ws.cell(row=2, column=7, value=None)
        wb.save(excel_path)
        wb.close()
        cfg = {"excel_path": str(excel_path), "sheet_name": "Daily Expenses"}
        mapping = read_all_categories(cfg)
        assert len(mapping) == 0
