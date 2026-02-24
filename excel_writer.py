from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment


class ExcelError(Exception):
    pass


def read_all_categories(config: dict) -> dict:
    """
    Read all rows from the Excel sheet and return an item→category mapping.

    Used by the --seed command to bootstrap categories.json.
    Returns an empty dict if the file/sheet is missing or empty.
    """
    excel_path = config["excel_path"]
    sheet_name = config["sheet_name"]

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
    except (FileNotFoundError, PermissionError):
        return {}

    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_col=5, max_col=6, values_only=True))
    wb.close()

    mapping = {}
    for category, item in rows:
        if item and category:
            mapping[str(item)] = str(category)

    return mapping


def append_row(config: dict, transaction: dict) -> None:
    """
    Append a transaction row to the configured Excel sheet.

    Columns written: Year (A), Month (B), Date (C), Amount (D), Category (E), Item (F)

    Raises:
        ExcelError: if the file or sheet cannot be found/written.
    """
    excel_path = config["excel_path"]
    sheet_name = config["sheet_name"]

    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        raise ExcelError(
            f"Excel file not found at: {excel_path}\n"
            "Check the 'excel_path' value in config.json."
        )
    except PermissionError:
        raise ExcelError(
            f"Cannot open the Excel file — it may be open in another application.\n"
            f"Close {excel_path} and try again."
        )

    if sheet_name not in wb.sheetnames:
        available = ", ".join(f'"{s}"' for s in wb.sheetnames)
        raise ExcelError(
            f"Sheet \"{sheet_name}\" not found in the workbook.\n"
            f"Available sheets: {available}\n"
            "Check the 'sheet_name' value in config.json."
        )

    ws = wb[sheet_name]
    next_row = ws.max_row + 1

    # Parse date to extract year and month
    dt = datetime.strptime(transaction["date"], "%m/%d/%Y")
    year = dt.year
    month = dt.strftime("%b")           # e.g. "Feb"

    AMOUNT_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
    CENTER = Alignment(horizontal="center")
    RIGHT = Alignment(horizontal="right")

    for col, value in ((1, year), (2, month)):              # A: Year, B: Month
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = CENTER
    date_cell = ws.cell(row=next_row, column=3, value=dt)   # C: Date
    date_cell.number_format = "mm-dd-yy"
    date_cell.alignment = CENTER
    amt_cell = ws.cell(row=next_row, column=4, value=transaction["amount"])  # D: Amount
    amt_cell.number_format = AMOUNT_FMT
    amt_cell.alignment = CENTER
    cat_cell = ws.cell(row=next_row, column=5, value=transaction["category"])  # E: Category
    cat_cell.alignment = RIGHT
    item_cell = ws.cell(row=next_row, column=6, value=transaction["item"])     # F: Item
    item_cell.alignment = RIGHT

    try:
        wb.save(excel_path)
    except PermissionError:
        raise ExcelError(
            f"Cannot save the Excel file — it may be open in another application.\n"
            f"Close {excel_path} and try again."
        )
