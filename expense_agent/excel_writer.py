from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment


class ExcelError(Exception):
    pass


def _open_workbook(config, read_only=False):
    """Open the configured workbook and return (wb, ws, excel_path).

    Raises ExcelError if the file or sheet cannot be found/opened.
    """
    excel_path = config["excel_path"]
    sheet_name = config["sheet_name"]

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=read_only)
    except FileNotFoundError:
        raise ExcelError(
            f"Excel file not found at: {excel_path}\n"
            "Check the 'excel_path' value in config.json."
        )
    except PermissionError:
        raise ExcelError(
            f"Cannot open the Excel file — it may be open in another application.\n"
            f"Close {excel_path} and try again."
        )

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ExcelError(
            f"Sheet \"{sheet_name}\" not found in the workbook.\n"
            f"Available sheets: {', '.join(wb.sheetnames)}"
        )

    return wb, wb[sheet_name], excel_path


def _save_workbook(wb, excel_path):
    """Save the workbook, raising ExcelError on permission issues."""
    try:
        wb.save(excel_path)
    except PermissionError:
        raise ExcelError(
            f"Cannot save the Excel file — it may be open in another application.\n"
            f"Close {excel_path} and try again."
        )


def read_all_categories(config: dict) -> dict:
    """
    Read all rows from the Excel sheet and return an item→subcategory mapping.

    Used by the --seed command to bootstrap categories.json.
    Returns an empty dict if the file/sheet is missing or empty.
    """
    try:
        wb, ws, _ = _open_workbook(config, read_only=True)
    except ExcelError:
        return {}

    rows = list(ws.iter_rows(min_col=6, max_col=7, values_only=True))
    wb.close()

    mapping = {}
    for subcategory, item in rows:
        if item and subcategory:
            mapping[str(item)] = str(subcategory)

    return mapping


def append_row(config: dict, transaction: dict) -> None:
    """
    Append a transaction row to the configured Excel sheet.

    Columns written: Year (A), Month (B), Date (C), Amount (D),
                     Category (E), Subcategory (F), Item (G)

    Raises:
        ExcelError: if the file or sheet cannot be found/written.
    """
    wb, ws, excel_path = _open_workbook(config)
    next_row = ws.max_row + 1

    # Parse date to extract year and month
    dt = datetime.strptime(transaction["date"], "%m/%d/%Y")
    year = dt.year
    month = dt.strftime("%b")           # e.g. "Feb"

    AMOUNT_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
    CENTER = Alignment(horizontal="center")
    RIGHT = Alignment(horizontal="right")

    for col, value in ((1, year), (2, month)):              # A: Year, B: Month
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = CENTER
    date_cell = ws.cell(row=next_row, column=3, value=dt)   # C: Date
    date_cell.number_format = "mm-dd-yy"
    date_cell.alignment = CENTER
    amt_cell = ws.cell(row=next_row, column=4, value=transaction["amount"])  # D: Amount
    amt_cell.number_format = AMOUNT_FMT
    amt_cell.alignment = CENTER
    cat_cell = ws.cell(row=next_row, column=5, value=transaction["category"])      # E: Category
    cat_cell.alignment = RIGHT
    subcat_cell = ws.cell(row=next_row, column=6, value=transaction["subcategory"])  # F: Subcategory
    subcat_cell.alignment = RIGHT
    item_cell = ws.cell(row=next_row, column=7, value=transaction["item"])          # G: Item
    item_cell.alignment = RIGHT

    _save_workbook(wb, excel_path)


def remove_last_row(config: dict) -> dict | None:
    """
    Remove the last data row from the Excel sheet.

    Returns the removed transaction as a dict, or None if the sheet is empty.

    Raises:
        ExcelError: if the file or sheet cannot be found/written.
    """
    wb, ws, excel_path = _open_workbook(config)
    last_row = ws.max_row

    if last_row < 2:
        wb.close()
        return None

    # Read the row before deleting
    date_val = ws.cell(row=last_row, column=3).value
    amount = ws.cell(row=last_row, column=4).value
    category = ws.cell(row=last_row, column=5).value
    subcategory = ws.cell(row=last_row, column=6).value
    item = ws.cell(row=last_row, column=7).value

    date_str = ""
    if isinstance(date_val, datetime):
        date_str = date_val.strftime("%m/%d/%Y")
    elif date_val:
        date_str = str(date_val)

    ws.delete_rows(last_row)
    _save_workbook(wb, excel_path)

    return {
        "date": date_str,
        "item": str(item) if item else "",
        "category": str(category) if category else "",
        "subcategory": str(subcategory) if subcategory else "",
        "amount": float(amount) if amount else 0.0,
    }
